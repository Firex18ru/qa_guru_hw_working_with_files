from zipfile import ZipFile
from pypdf import PdfReader
from io import TextIOWrapper, BytesIO
from openpyxl import load_workbook
import csv


def test_pdf(test_zip_folder):
    with ZipFile("resources/folder.zip", "r") as folder_zip:
        with folder_zip.open("tmp/Test_pdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = len(reader.pages)
            second_page_found = reader.pages[1]
            text = second_page_found.extract_text()
            expected_text = "Class aptent taciti sociosqu ad litora"
            assert expected_text in text
            assert page == 3


def test_csv(test_zip_folder):
    with ZipFile("resources/folder.zip", "r") as folder_zip:
        with folder_zip.open("tmp/Test_csv.csv") as csv_file:
            reader = csv.reader(TextIOWrapper(csv_file, encoding="utf-8"))
            expected_row = ["1", "Dett", "Male", "18", "21/05/2015", "Great Britain"]
            assert any(expected_row == row for row in reader)


def test_xlsx(test_zip_folder):
    with ZipFile("resources/folder.zip", "r") as folder_zip:
        with folder_zip.open("tmp/Test_xlsx.xlsx") as xlsx_file:
            workbook = load_workbook(BytesIO(xlsx_file.read()))
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            expected_row = (1, "Dett", "Male", 18, "21/05/2015", "Great Britain")
            assert expected_row in data
